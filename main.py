from quicksort_hoare import quicksort as quicksort_hoare
from quicksort_lomuto import quicksort as quicksort_lomuto
from heapsort import heap_sort
from mergesort import merge_sort

import time
import statistics
from tabulate import tabulate
import csv
import pandas as pd
import os
import openpyxl

N = 500 #alterar para o tamanho do arquivo
FILENAME = f'numbers{N}.txt' #alterar para o arquivo desejado
REPS = 5

def save_to_csv(data, filename, headers):
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data)

def read_file(filename):   #alterar para ler um arquivo de numeros
    file_path = f'lists/{filename}'
    with open(file_path, 'r', encoding='utf-8') as f:
        return [int(line.strip()) for line in f.readlines()]

def main():
    algorithms = {
        "Quicksort (Hoare)": quicksort_hoare,
        "Quicksort (Lomuto)": quicksort_lomuto,
        "Heapsort": heap_sort,
        "Mergesort": merge_sort
    }

    results = {name: [] for name in algorithms}

    for _ in range(REPS):
        numbers = read_file(FILENAME)
        for name, algorithm in algorithms.items():
            numbers_copy = numbers.copy()
            start_time = time.time()
            if name == "Heapsort":
                algorithm(numbers_copy)
            else:
                algorithm(numbers_copy, 0, len(numbers_copy) - 1)
            end_time = time.time()
            elapsed_time = (end_time - start_time) * 1000  # Convert to milliseconds
            results[name].append(elapsed_time)

    # Exibir resultados detalhados
    detailed_table = []
    for name, tempos in results.items():
        for i, tempo in enumerate(tempos):
            detailed_table.append([name, i + 1, f"{tempo:.2f} ms"])

    print(tabulate(detailed_table, headers=["Algoritmo", "Execução", "Tempo (ms)"]))

    # Exibir resultados
    table = []
    for name, tempos in results.items():
        avg_time = statistics.mean(tempos)
        std_dev = statistics.stdev(tempos)
        table.append([name, f"{avg_time:.2f} ms", f"{std_dev:.2f} ms"])

    print(tabulate(table, headers=["Algoritmo", "Tempo Médio", "Desvio Padrão"]))

    # Criar diretórios se não existirem
    if not os.path.exists('results'):
        os.makedirs('results')

    def save_to_excel(data, filename, sheet_name, headers):
        if not os.path.exists(filename):
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
        else:
            workbook = openpyxl.load_workbook(filename)
        
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(headers)
        for row in data:
            sheet.append(row)
        
        workbook.save(filename)

    # Salvar resultados detalhados em Excel
    for name, tempos in results.items():
        individual_data = [[i + 1, f"{time:.4f} ms"] for i, time in enumerate(tempos)]
        avg_time = statistics.mean(tempos)
        std_dev = statistics.stdev(tempos)
        individual_data.append(["Média", f"{avg_time:.4f} ms"])
        individual_data.append(["Desvio Padrão", f"{std_dev:.4f} ms"])
        save_to_excel(individual_data, f"results/{name.replace(' ', '_').lower()}_details.xlsx", f"N={N}", ["Execução", "Tempo (ms)"])
    
    # Salvar resultados comparativos em Excel
    comparative_data = [[name, f"{statistics.mean(tempos):.4f} ms"] for name, tempos in results.items()]
    save_to_excel(comparative_data, f"results/comparative_results.xlsx", f"N={N}", ["Algoritmo", "Tempo Médio (ms)"])

if __name__ == "__main__":
    main()